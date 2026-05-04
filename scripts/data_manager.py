import asyncio
import aiosqlite
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from datetime import datetime

class DataManager:
    def __init__(self, db_name="test_results.db", excel_name="test_results.xlsx"):
        self.db_name = db_name
        self.excel_name = excel_name
        self.queue = asyncio.Queue()
        self.worker_task = None
        self.wb = openpyxl.Workbook()
        
        # Настройка цветов
        self.fills = {
            "Успех": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), # Светло-зеленый
            "Ошибка": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), # Светло-красный
            "Таймаут": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), # Желтый
            "Предупреждение": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "Комментарий": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        }

        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

    async def init_db(self):
        async with aiosqlite.connect(self.db_name) as db:
            await db.execute('''
                CREATE TABLE IF NOT EXISTS test_results (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT,
                    test_source TEXT,
                    command TEXT,
                    target_value TEXT,
                    status TEXT,
                    output TEXT
                )
            ''')
            await db.commit()

    async def start(self):
        self.loop = asyncio.get_running_loop()
        await self.init_db()
        self.worker_task = asyncio.create_task(self._worker())

    async def stop(self):
        if self.worker_task:
            await self.queue.join()
            self.worker_task.cancel()
            
            # АВТО-ВЫРАВНИВАНИЕ СТОЛБЦОВ ПЕРЕД СОХРАНЕНИЕМ
            for sheet in self.wb.worksheets:
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                length = len(str(cell.value))
                                if length > max_length:
                                    max_length = length
                        except: pass
                    adjusted_width = min(max_length + 2, 50) # Ограничим до 50, чтобы не было слишком широко
                    sheet.column_dimensions[column].width = adjusted_width

            try:
                await asyncio.to_thread(self.wb.save, self.excel_name)
                print(f"✓ Отчет сохранен: {self.excel_name}")
            except Exception as e:
                print(f"✗ Ошибка сохранения Excel: {e}")

    def sync_add_result(self, test_source, command, status, output, target_value="N/A"):
        """Потокобезопасная функция"""
        if hasattr(self, 'loop') and self.loop.is_running():
            asyncio.run_coroutine_threadsafe(
                self.add_result(test_source, command, status, output, target_value), 
                self.loop
            )

    async def add_result(self, test_source, command, status, output, target_value="N/A"):
        safe_output = ILLEGAL_CHARACTERS_RE.sub(r'', str(output))
        safe_command = ILLEGAL_CHARACTERS_RE.sub(r'', str(command))
        
        result_data = {
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "test_source": test_source,
            "command": safe_command,
            "target_value": str(target_value),
            "status": status,
            "output": safe_output
        }
        await self.queue.put(result_data)

    async def _worker(self):
        try:
            async with aiosqlite.connect(self.db_name) as db:
                while True:
                    item = await self.queue.get()
                    try:
                        # в БД
                        await db.execute(
                            "INSERT INTO test_results (timestamp, test_source, command, target_value, status, output) VALUES (?, ?, ?, ?, ?, ?)",
                            (item['timestamp'], item['test_source'], item['command'], item['target_value'], item['status'], item['output'])
                        )
                        await db.commit()

                        # excel
                        sheet_name = item['test_source'][:31]
                        if sheet_name not in self.wb.sheetnames:
                            ws = self.wb.create_sheet(title=sheet_name)
                            ws.append(["Время", "Источник", "Команда", "Ожидаемое значение", "Статус", "Вывод"])
                            for cell in ws[1]:
                                cell.font = Font(bold=True)
                        else:
                            ws = self.wb[sheet_name]
                        
                        # Добавляет сроку
                        row_data = [item['timestamp'], item['test_source'], item['command'], item['target_value'], item['status'], item['output'][:1000]]
                        ws.append(row_data)

                        last_row = ws.max_row
                        status_cell = ws.cell(row=last_row, column=5)
                        if item['status'] in self.fills:
                            status_cell.fill = self.fills[item['status']]

                    except Exception as e:
                        print(f"Ошибка записи строки: {e}")
                    finally:
                        self.queue.task_done()
        except asyncio.CancelledError:
            pass