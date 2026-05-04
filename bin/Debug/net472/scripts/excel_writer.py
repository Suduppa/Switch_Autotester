import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re
import unicodedata


class ExcelTestWriter:
    def __init__(self, filename="test_results.xlsx"):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.test_sheets = {}
        self.current_sheet = None
        self.row_counter = 1

        # Стили для форматирования
        self.styles = {
            "header": Font(bold=True, color="FFFFFF"),
            "header_fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            "success": Font(color="00B050"),
            "error": Font(color="FF0000"),
            "warning": Font(color="FFC000"),
            "comment": Font(color="808080", italic=True),
            "centered": Alignment(horizontal="center", vertical="center"),
            "left": Alignment(horizontal="left", vertical="center", wrap_text=True)
        }

    def clean_string_for_excel(self, text):
        """Очищает строку от недопустимых символов для Excel"""
        if text is None:
            return ""

        if not isinstance(text, str):
            text = str(text)

        cleaned = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)

        replacements = {
            '\u0000': '',  # Null
            '\u0001': '',  # Start of Heading
            '\u0002': '',  # Start of Text
            '\u0003': '',  # End of Text
            '\u0004': '',  # End of Transmission
            '\u0005': '',  # Enquiry
            '\u0006': '',  # Acknowledge
            '\u0007': '',  # Bell
            '\u0008': '',  # Backspace
            '\u000B': '\n',  # Vertical Tab -> перевод строки
            '\u000C': '\n',  # Form Feed -> перевод строки
            '\u000E': '',  # Shift Out
            '\u000F': '',  # Shift In
            '\u0010': '',  # Data Link Escape
            '\u0011': '',  # Device Control 1
            '\u0012': '',  # Device Control 2
            '\u0013': '',  # Device Control 3
            '\u0014': '',  # Device Control 4
            '\u0015': '',  # Negative Acknowledge
            '\u0016': '',  # Synchronous Idle
            '\u0017': '',  # End of Transmission Block
            '\u0018': '',  # Cancel
            '\u0019': '',  # End of Medium
            '\u001A': '',  # Substitute
            '\u001B': '',  # Escape
            '\u001C': '',  # File Separator
            '\u001D': '',  # Group Separator
            '\u001E': '',  # Record Separator
            '\u001F': '',  # Unit Separator
            '\u007F': '',  # Delete
        }

        for old, new in replacements.items():
            cleaned = cleaned.replace(old, new)

        # Нормализуем Unicode символы
        cleaned = unicodedata.normalize('NFKD', cleaned)

        return cleaned.strip()

    def create_sheet_for_test(self, test_name):
        """Создает новый лист для теста"""
        safe_sheet_name = self.clean_string_for_excel(test_name)[:31]

        if safe_sheet_name in self.test_sheets:
            ws = self.test_sheets[safe_sheet_name]
            self.wb.remove(ws)

        try:
            ws = self.wb.create_sheet(title=safe_sheet_name)
        except:
            ws = self.wb.create_sheet(title=f"Test_{len(self.test_sheets) + 1}")

        self.test_sheets[test_name] = ws
        self.current_sheet = ws
        self.row_counter = 1

        self._write_headers()
        return ws

    def _write_headers(self):
        """Записывает заголовки столбцов"""
        if not self.current_sheet:
            return

        headers = ["№ теста", "Значение", "Команда", "Статус"]

        for col_idx, header in enumerate(headers, 1):
            cell = self.current_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles["header"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["centered"]
            # Автоподбор ширины столбца
            self.current_sheet.column_dimensions[get_column_letter(col_idx)].width = 25

    def write_test_result(self, test_index, value, command, status, output=""):
        """Записывает результат одного теста"""
        if not self.current_sheet:
            return

        self.row_counter += 1
        row = self.row_counter

        safe_value = self.clean_string_for_excel(value)
        safe_command = self.clean_string_for_excel(command)
        safe_output = self.clean_string_for_excel(output)
        safe_status = self.clean_string_for_excel(status)

        self.current_sheet.cell(row=row, column=1, value=test_index)
        self.current_sheet.cell(row=row, column=2, value=safe_value)

        if safe_output:
            full_command = f"{safe_command}\n\nВывод:\n{safe_output}"
        else:
            full_command = safe_command

        self.current_sheet.cell(row=row, column=3, value=full_command)
        self.current_sheet.cell(row=row, column=3).alignment = self.styles["left"]

        status_cell = self.current_sheet.cell(row=row, column=4, value=safe_status)

        status_lower = safe_status.lower()
        if "success" in status_lower or "успех" in status_lower:
            status_cell.font = self.styles["success"]
        elif any(word in status_lower for word in ["error", "ошибка", "failed", "invalid", "syntax"]):
            status_cell.font = self.styles["error"]
        elif "warning" in status_lower or "предупреждение" in status_lower:
            status_cell.font = self.styles["warning"]
        elif "comment" in status_lower or "комментарий" in status_lower:
            status_cell.font = self.styles["comment"]

        self.current_sheet.row_dimensions[row].height = 60

    def write_comment(self, comment):
        """Записывает комментарий"""
        if not self.current_sheet:
            return

        self.row_counter += 1
        row = self.row_counter

        safe_comment = self.clean_string_for_excel(comment)
        cell = self.current_sheet.cell(row=row, column=3, value=safe_comment)
        cell.font = self.styles["comment"]
        cell.alignment = self.styles["left"]

        self.current_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    def save(self):
        """Сохраняет файл Excel"""
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        self.wb.properties.title = "Результаты автотестирования"
        self.wb.properties.subject = f"Тесты от {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        try:
            self.wb.save(self.filename)
            print(f"Результаты сохранены в файл: {self.filename}")
        except Exception as e:
            print(f"Ошибка сохранения файла: {e}")

            backup_name = f"test_results_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.wb.save(backup_name)
            print(f"Результаты сохранены в резервный файл: {backup_name}")